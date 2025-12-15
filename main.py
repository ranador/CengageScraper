import os
import sys

from dataclasses import dataclass

from PyQt5.QtWidgets import QApplication, QMainWindow, QWidget, QMessageBox, QLabel, QListWidget, QPushButton, QVBoxLayout, QHBoxLayout, QFileDialog, QTextEdit, QLineEdit, QGroupBox
from PyQt5.QtGui import QIcon
from PyQt5.QtCore import QSize

import openpyxl
from openpyxl.styles import PatternFill, Alignment, Font
from openpyxl.styles.borders import Border, Side
from openpyxl.utils.cell import get_column_letter

import pandas as pd
import pickle
import csv
import re
from pathlib import Path

@dataclass
class Settings:
    commentCode: str = None
    documentationCode: str = None
    courseNumber: str = None
    roster: pd.DataFrame = None
    header_length: int = 9
    asst_name_idx: int = 4
    asst_points_idx: int = 7
    
    output_directory: str = None

    def save(self, filepath='settings/settings.pkl'):
        with open(filepath, 'wb') as f:
            pickle.dump(self, f)

    @classmethod
    def load(cls, filepath='settings/settings.pkl'):
        try:
            with open(filepath, 'rb') as f:
                return pickle.load(f)
        except FileNotFoundError:
            return cls()

@dataclass
class Data:
    raw_data: list = None
    header_data: list = None
    final_data: pd.DataFrame = None
    name: str = None
    points: list[float] = None
    n_questions: int = None
    qCodes: list[tuple[str, int]] = None
    comment_idx: int = None
    documentation_idx: int = None

class MainWindow(QMainWindow):
    def __init__(self):
        # Ensures all initialization code from the inherited class is executed
        super().__init__()

        # Set application spanning styles
        btnMargin = 25
        btnIconHeight = 32
        btnIconSize = QSize(btnIconHeight, btnIconHeight)
        btnHeight = btnIconHeight + btnMargin

        self.setStyleSheet(f"""
            QGroupBox QPushButton {{
                height:  {btnHeight}              
            }}
            QLabel {{
                height: 100px;
                font-family: Arial;
            }}
            QListWidget, QTextEdit, QLineEdit {{
                border: 1px solid grey;
            }}
            QGroupBox {{
                min-width: 200px;
                min-height: 400px;
            }}
        """)

        # Global Variables
        self.settings = Settings().load()
        self.asst_data = None

        # Setup output directory
        self.output_dir = 'output'
        if not os.path.exists(self.output_dir):
            os.makedirs(self.output_dir)

        # Initialize main widget
        main_widget = QWidget()
        self.setCentralWidget(main_widget)
        self.setWindowTitle("Cengage Data Scraper")

        # Create the layout
        main_layout = QHBoxLayout()

        # Settings Group
        grpOptions = QGroupBox("Settings")
        optionLayout = QVBoxLayout()
        grpOptions.setLayout(optionLayout)

        lblCourseNumber = QLabel("Course Number")
        self.lnEdtCourseNumber = QLineEdit(f'{self.settings.courseNumber if self.settings.courseNumber else ""}')
        lblDocumentationCode = QLabel("Documentation Code")
        self.lnEdtDocumentationCode = QLineEdit(f'{self.settings.documentationCode if self.settings.documentationCode else ""}')
        lblCommentCode = QLabel("Comment Code")
        self.lnEdtCommentCode = QLineEdit(f'{self.settings.commentCode if self.settings.commentCode else ""}')
        self.btnSaveSettings = QPushButton("Save Settings", enabled=True, clicked=self.save_settings)
        
        optionLayout.addWidget(lblCourseNumber)
        optionLayout.addWidget(self.lnEdtCourseNumber)
        optionLayout.addWidget(lblDocumentationCode)
        optionLayout.addWidget(self.lnEdtDocumentationCode)
        optionLayout.addWidget(lblCommentCode)
        optionLayout.addWidget(self.lnEdtCommentCode)
        optionLayout.addStretch()
        optionLayout.addWidget(self.btnSaveSettings)

        # Roster Group
        grpRoster = QGroupBox("Current Roster")
        rosterLayout = QVBoxLayout()
        grpRoster.setLayout(rosterLayout)

        self.listRoster = QListWidget()
        self.btnRoster = QPushButton("Load Roster", enabled=True)

        rosterLayout.addWidget(self.listRoster)
        rosterLayout.addWidget(self.btnRoster)
        
        # Data Group
        grpData = QGroupBox("Data")
        dataLayout = QVBoxLayout()
        grpData.setLayout(dataLayout)

        self.txtData = QTextEdit()
        self.txtData.setMinimumWidth(500)
        self.txtData.setLineWrapMode(QTextEdit.NoWrap)
        self.btnLoadData = QPushButton("Load Data", enabled=True, clicked=self.load_data_file)
        self.btnExportData = QPushButton("Export Data", enabled=False, clicked=self.export)
        self.btnLoadData.setIcon(QIcon('resources/icons/parse.ico'))
        self.btnLoadData.setIconSize(btnIconSize)
        self.btnExportData.setIcon(QIcon('resources/icons/export.png'))
        self.btnExportData.setIconSize(btnIconSize)

        dataLayout.addWidget(self.txtData)
        tempHLayout = QHBoxLayout()
        tempHLayout.addWidget(self.btnLoadData)
        tempHLayout.addWidget(self.btnExportData)
        dataLayout.addLayout(tempHLayout)

        # Connect buttons to functions
        self.btnRoster.clicked.connect(self.setup_roster)

        # Add widgets to the grid
        main_layout.addWidget(grpOptions, stretch=1)
        main_layout.addWidget(grpRoster, stretch=1)
        main_layout.addWidget(grpData, stretch=3)

        # Set the layout for the GUI
        main_widget.setLayout(main_layout)

        self.populate_roster()

    def show_message(self, message):
        msg = QMessageBox()
        msg.setIcon(QMessageBox.Information)
        msg.setText(message)
        msg.setWindowTitle("Message")
        msg.exec_()

    def show_error(self, message):
        msg = QMessageBox()
        msg.setIcon(QMessageBox.Critical)
        msg.setText(message)
        msg.setWindowTitle("Error")
        msg.exec_()

    def confirm_action(self, message):
        reply = QMessageBox.question(self, 'Confirmation', message, QMessageBox.Yes | QMessageBox.No, QMessageBox.No)

        if reply == QMessageBox.Yes:
            return True
        else:
            return False

    def open_file_dialog(self, extensions):
        options = QFileDialog.Options()
        file_name, _ = QFileDialog.getOpenFileName(self, "Open File", "", extensions, options=options)

        if file_name:
            return file_name

    def populate_roster(self):
        if self.settings.roster is None:
            self.show_message('You need to establish a class roster.')
            return
        for name, section in zip(self.settings.roster['Cadet Name'], self.settings.roster['Section']):
            self.listRoster.addItem(f'{name} ({section})')
        
        self.btnLoadData.setEnabled(True)
    
    def populate_data_view(self, asst_data: Data):
        self.txtData.setHtml(asst_data.final_data.to_html())
        self.btnExportData.setEnabled(True)

    def process_names(self, text):
        match = re.search(r',[^ ]+', text)
        text = text[:match.end()] if match else text

        return text

    def save_settings(self):
        self.settings.courseNumber = self.lnEdtCourseNumber.text() if self.lnEdtCourseNumber.text() != '' else None
        self.settings.commentCode = self.lnEdtCommentCode.text() if self.lnEdtCommentCode.text() != '' else None
        self.settings.documentationCode = self.lnEdtDocumentationCode.text() if self.lnEdtDocumentationCode.text() != '' else None
        
        self.settings.save()

    def setup_roster(self):
        # Verify Data
        if self.settings.courseNumber is None: 
            self.show_message('You must enter course information before continuing')
            return

        # Verify intent
        if self.settings.roster is not None and not self.confirm_action('This will reset the current roster, do you wish to continue?'): return

        # Open file dialog
        file_path = self.open_file_dialog("Excel Files (*.xlsx *.xls)")
        if not file_path: 
            return

        try:
            df = pd.read_excel(file_path, skiprows=1)

            df["Course Number"] = df["Course Number"].str.strip()
            df = df[df["Course Number"] == f'{self.settings.courseNumber}'][["Section", "Email", "Cadet Name"]]
            df["Cadet Name"] = df["Cadet Name"].str.strip().map(self.process_names)

            self.settings.roster = df
            self.settings.save()
            self.populate_roster()

        except Exception as e:
            self.show_error(f"Failed to load file\n{e}")
    
    def load_data_file(self):
        file_path = self.open_file_dialog("CSV Files (*.csv)")
        if not file_path:
            return
        
        with open(file_path, 'r') as f:
            data = f.readlines()
            data = [x.strip() for x in data]
            header = data[:self.settings.header_length]
            body = data[self.settings.header_length:]
        
        asst_data = self.process_header(Data(raw_data=body, header_data=header))
        asst_data = self.parse_data(asst_data)
        
        self.populate_data_view(asst_data)
        self.asst_data = asst_data

    def process_header(self, asst_data: Data):
        header = asst_data.header_data

        points = [float(x) for x in header[self.settings.asst_points_idx].split(',') if x != '' and x != 'Points']
        name = header[self.settings.asst_name_idx].split(',')[1]
        qCodes = [(x, j) for j, x in enumerate(header[6].split(',')) if x.isdecimal()]

        for code, j in qCodes:
            if code == self.settings.commentCode:
                asst_data.comment_idx = j
            if code == self.settings.documentationCode:
                asst_data.documentation_idx = j

        asst_data.name = name
        asst_data.points = points
        asst_data.n_questions = len(points)
        asst_data.qCodes = qCodes

        return asst_data
    
    def parse_data(self, data: Data):
        csvReader = csv.reader(data.raw_data)
        students: pd.DataFrame = self.settings.roster
        result = []
        email = ''
        name = ''
        comment = None
        documentation = None

        for i, row in enumerate(csvReader):
            if i%2 == 1:
                filter = students['Email'].isin([email])
                if filter.any():
                    section = students[filter]['Section'].iloc[0]
                    nRow = [name, email, section]
                    
                    for j, entry in enumerate(row):
                        if j == 3:
                            nRow.append(float(entry))
                        if j >= 4:
                            if float(entry) == 0:
                                nRow.append('-')
                            elif float(entry) == data.points[j - 4]:
                                nRow.append(1.0)
                            else:
                                nRow.append(0.5)
                    
                    if data.documentation_idx:
                        nRow[-1] = documentation if documentation else ''
                    if data.comment_idx:
                        nRow[-2 if data.documentation_idx else -1] = comment if comment else ''

                    result.append(nRow)          
            else:
                email = row[1].split('@usafa')[0]
                name = self.process_names(row[0])
                comment = row[data.comment_idx] if data.comment_idx else None
                documentation = row[data.documentation_idx] if data.documentation_idx else None

        header = ['Name','Email','Section', 'Total'] + [f'Q{x + 1}' for x in range(data.n_questions)]

        if data.documentation_idx:
            header[-1] = 'Documentation'
        if data.comment_idx:
            header[-2 if data.documentation_idx else -1] = 'Comment'

        data.final_data = pd.DataFrame(result, columns=header)

        return data

    def export(self):
        asst_data = self.asst_data
        output_dir = Path('output') / asst_data.name.strip('"')
        file_path = output_dir / 'output.xlsx'

        if not output_dir.exists():
            output_dir.mkdir()
        
        if file_path.exists():
            output_wb = openpyxl.load_workbook(file_path)
        else:
            output_wb = openpyxl.Workbook()
        
        for section in asst_data.final_data['Section'].unique():

            if output_wb.sheetnames[0] == 'Sheet':
                ws = output_wb.active
                ws.title = f'{section}'
            elif section in output_wb.sheetnames:
                ws = output_wb[section]
            else:
                ws = output_wb.create_sheet(title=section)

            self.generate_excel_table(asst_data.final_data[asst_data.final_data['Section'] == section], ws)
        
        output_wb.save(file_path)

    def _pixel_to_pt(self, x):
        return x / 7.0

    def _truncate_string(self, s, max_length=20):
        return s if len(s) <= max_length else s[:max_length] + '...'

    def _truncate_or_pad_string(self, s, max_length=70):
        return (s[:max_length - 3] + '...') if len(s) > max_length else s.ljust(max_length)

    def generate_excel_table(self, df: pd.DataFrame, ws):
        # define fill colors
        greenFill = PatternFill(start_color='FF00B050', end_color='FF00B050', fill_type='solid')
        redFill = PatternFill(start_color='FFC00000', end_color='FFC00000', fill_type='solid')
        borderFill = PatternFill(start_color='FF808080', end_color='FF808080', fill_type='solid')
        headerFill = PatternFill(start_color='FFD9D9D9', end_color='FFD9D9D9', fill_type='solid')
        whiteFill = PatternFill(start_color='FFFFFFFF', end_color='FFFFFFFF', fill_type='solid')
        warningFill = PatternFill(start_color='FFF59412', end_color='FFF59412', fill_type='solid')

        # Define border styles
        thin_all_sides = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        thin_bottom = Border(bottom=Side(style='thin'))
        thin_bottom_sides = Border(left=Side(style='thin'), right=Side(style='thin'), bottom=Side(style='thin'))
        thin_sides = Border(left=Side(style='thin'), right=Side(style='thin'))

        df = df.drop(['Email', 'Section'], axis=1, inplace=False).reset_index(drop=True)
        question_cols = [col for col in df.columns if col.startswith('Q')]
        n_questions = len(question_cols)
        n_students = len(df)
        documentation = 'Documentation' in df.columns
        comment = 'Comment' in df.columns
        last_col = 0
        pasteLoc = None

        center_align = Alignment(horizontal='center', vertical='center')
        left_align_indent = Alignment(horizontal='left', vertical='center', indent=1)

        # Set column widths
        ws.column_dimensions['A'].width = self._pixel_to_pt(22)
        ws.column_dimensions['B'].width = self._pixel_to_pt(200)
        ws.column_dimensions['C'].width = self._pixel_to_pt(48)

        for i in range(n_questions):
            last_col = 4 + i
            col = get_column_letter(last_col)
            ws.column_dimensions[col].width = self._pixel_to_pt(30)
        
        if comment:
            last_col += 1
            col = get_column_letter(last_col)
            ws.column_dimensions[col].width = self._pixel_to_pt(665)
            pasteLoc = (col, len(df['Name']) + 8)

        if documentation:
            last_col += 1
            col = get_column_letter(last_col)
            ws.column_dimensions[col].width = self._pixel_to_pt(294)
        
        ws.column_dimensions[get_column_letter(last_col + 1)].width = self._pixel_to_pt(21)

        # Set row heights
        for i in range(1, n_students + 21):
            ws.row_dimensions[i].height = 20
        
        for col in range(1, len(df.columns) + 2):
            for row in range(1, len(df['Name']) + 6):
                ws[f'{get_column_letter(col)}{row}'].fill = borderFill

        # Write Title
        ws['B2'] = self.asst_data.name.strip('"')
        ws['B2'].fill = headerFill
        ws['B2'].border = thin_all_sides
        ws.merge_cells(f'B2:{get_column_letter(last_col)}2')
        ws['B2'].alignment = center_align

        # Write Headers
        for i, title in enumerate(df.columns):
            cell = f'{get_column_letter(2 + i)}4'
            title = title if title != 'Comment' else 'What did you find interesting/useful/confusing?'
            title = title if title != 'Documentation' else 'Documentation Statement'
            ws[cell] = title
            ws[cell].alignment = center_align if title.startswith(('Q', 'T')) else left_align_indent
            ws[cell].fill = headerFill
            ws[cell].border = thin_all_sides

        # Write Student Data
        for idx, student in df.iterrows():
            col = 2
            row = idx + 5

            ws[f'{get_column_letter(col)}{row}'] = f'{student["Name"]}'
            ws[f'{get_column_letter(col)}{row}'].alignment = left_align_indent
            ws[f'{get_column_letter(col)}{row}'].fill = whiteFill
            ws[f'{get_column_letter(col)}{row}'].border = thin_all_sides

            col += 1
            ws[f'{get_column_letter(col)}{row}'] = f'{student["Total"]}'
            ws[f'{get_column_letter(col)}{row}'].alignment = center_align
            ws[f'{get_column_letter(col)}{row}'].fill = whiteFill
            ws[f'{get_column_letter(col)}{row}'].border = thin_all_sides

            for i in range(n_questions):
                col = 4 + i
                cell = f'{get_column_letter(col)}{row}'
                val = student[f'Q{i + 1}']
                ws[cell] = val if val == '-' else ''
                ws[cell].alignment = center_align
                ws[cell].border = thin_all_sides
                if val == 1:
                    ws[cell].fill = greenFill
                elif val == 0.5:
                    ws[cell].fill = redFill
                elif val == '-':
                    ws[cell].fill = whiteFill
                else:
                    ws[cell].fill = warningFill
            
            if comment:
                col += 1
                ws[f'{get_column_letter(col)}{row}'] = self._truncate_or_pad_string(f'{student["Comment"]}', max_length=95)
                ws[f'{get_column_letter(col)}{row}'].alignment = left_align_indent
                ws[f'{get_column_letter(col)}{row}'].fill = whiteFill
                ws[f'{get_column_letter(col)}{row}'].border = thin_all_sides

            if documentation:
                col += 1
                ws[f'{get_column_letter(col)}{row}'] = self._truncate_or_pad_string(f'{student["Documentation"]}', max_length=40)
                ws[f'{get_column_letter(col)}{row}'].alignment = left_align_indent
                ws[f'{get_column_letter(col)}{row}'].fill = whiteFill
                ws[f'{get_column_letter(col)}{row}'].border = thin_all_sides

        if pasteLoc:
            filter_list = ['', ' ', '.', 'none', 'n/a', 'nope', 'negative', 'nothing yet', 'nothing', 'nothing.', 'nothing so far', 'none so far', 'none for now']
            allowed_comments = df[~df['Comment'].str.lower().isin(filter_list)]['Comment'].reset_index(drop=True)

            col, row = pasteLoc

            ws[f'{col}{row}'] = 'Copy and Paste Comments:'
            ws[f'{col}{row}'].border = thin_bottom

            for idx, student_comment in enumerate(allowed_comments.to_list()):
                ws[f'{col}{row + idx + 1}'] = student_comment
                ws[f'{col}{row + idx + 1}'].border = thin_sides if idx != len(allowed_comments) - 1 else thin_bottom_sides

        print(f'{question_cols=}, {documentation=}, {comment=}')

if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = MainWindow()
    window.show()
    sys.exit(app.exec_())